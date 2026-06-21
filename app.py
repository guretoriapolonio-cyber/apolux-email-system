"""
Apolux Technology & AI — Sistema de Disparo de E-mails v8.0
Backend Flask + Frontend moderno (HTML/CSS/JS)
Roda como app desktop (pywebview) OU como site local (navegador)
"""

from flask import Flask, render_template, request, jsonify, send_file
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import json
import os
import csv
import io
import time
import random
import uuid
import threading
import webbrowser
from datetime import datetime, timedelta
from dotenv import load_dotenv

load_dotenv()
GMAIL_USER = os.getenv("GMAIL_USER", "")
GMAIL_PASS = os.getenv("GMAIL_PASS", "")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

F_CONTATOS  = os.path.join(DATA_DIR, "contatos.json")
F_TEMPLATES = os.path.join(DATA_DIR, "templates.json")
F_HISTORICO = os.path.join(DATA_DIR, "historico.json")
F_BLACKLIST = os.path.join(DATA_DIR, "blacklist.json")
F_FOLLOWUP  = os.path.join(DATA_DIR, "followup.json")
F_CAMPANHAS = os.path.join(DATA_DIR, "campanhas.json")
F_AGENDADOS = os.path.join(DATA_DIR, "agendados.json")
F_ABERTURAS = os.path.join(DATA_DIR, "aberturas.json")
F_WPP       = os.path.join(DATA_DIR, "whatsapp.json")

TEMPLATES_PADRAO = [
    {"id": 1, "nome": "Clínica Estética", "tipo": "clinica",
     "assunto": "Reduzindo faltas e retrabalho na [Empresa]",
     "corpo": "Olá, [Nome],\n\nTudo bem?\n\nSou da Apolux, especializada em automação com IA para clínicas estéticas.\n\nAnalisei o perfil da [Empresa] e percebi que processos manuais como confirmação de agenda e follow-up podem estar custando clientes todo dia.\n\nO que resolvemos:\n→ Agendamento automático via WhatsApp com IA\n→ Lembretes 48h, 24h e 2h antes do procedimento\n→ Avaliação no Google automática pós-atendimento\n→ Recuperação de clientes inativos\n→ Dashboard com ROI em tempo real\n\nClínicas reduziram faltas em até 30% no primeiro mês.\n\nPosso mostrar em 15 minutos essa semana?\n\nAtt,\nApolux Technology & AI\n📲 (11) 99325-3806\n📧 apoluxai@gmail.com"},
    {"id": 2, "nome": "Escritório Contábil", "tipo": "escritorio",
     "assunto": "Automatizando tarefas repetitivas no seu escritório",
     "corpo": "Olá, [Nome],\n\nTudo bem?\n\nSou da Apolux, especializada em automação com IA para escritórios contábeis.\n\nAnalisei escritórios como o [Empresa] e percebi que boa parte do tempo vai para tarefas automatizáveis.\n\nO que resolvemos:\n→ Leitura automática de notas fiscais\n→ Cobrança mensal de documentos via WhatsApp\n→ Alertas de vencimentos fiscais\n→ IA para dúvidas frequentes dos clientes\n→ Dashboard com indicadores em tempo real\n\nPosso mostrar em 15 minutos essa semana?\n\nAtt,\nApolux Technology & AI\n📲 (11) 99325-3806\n📧 apoluxai@gmail.com"},
]

# ──────────────────────────────────────────────
# PERSISTÊNCIA
# ──────────────────────────────────────────────

def carregar(arq, pad):
    if os.path.exists(arq):
        with open(arq, "r", encoding="utf-8") as f:
            return json.load(f)
    return pad

def salvar(arq, dados):
    with open(arq, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def get_contatos():  return carregar(F_CONTATOS, [])
def get_templates():
    t = carregar(F_TEMPLATES, None)
    if t is None:
        salvar(F_TEMPLATES, TEMPLATES_PADRAO)
        return TEMPLATES_PADRAO
    return t
def get_historico(): return carregar(F_HISTORICO, [])
def get_blacklist(): return carregar(F_BLACKLIST, [])
def get_followup():  return carregar(F_FOLLOWUP, [])
def get_campanhas(): return carregar(F_CAMPANHAS, [])
def get_agendados(): return carregar(F_AGENDADOS, [])
def get_aberturas(): return carregar(F_ABERTURAS, {})
def get_wpp():        return carregar(F_WPP, [])

def proximo_id(lista): return max((t["id"] for t in lista), default=0) + 1

# ──────────────────────────────────────────────
# ENVIO DE E-MAIL
# ──────────────────────────────────────────────

def personalizar(texto, nome, empresa):
    return texto.replace("[Nome]", nome).replace("[Empresa]", empresa)

def enviar_email(destinatario, assunto, corpo, tracking_id=None):
    if not GMAIL_USER or not GMAIL_PASS:
        raise ValueError("Credenciais não configuradas no .env")
    msg = MIMEMultipart("alternative")
    msg["From"] = GMAIL_USER
    msg["To"] = destinatario
    msg["Subject"] = assunto
    corpo_html = corpo.replace("\n", "<br>")
    if tracking_id:
        corpo_html += f'<img src="http://localhost:5050/track/{tracking_id}" width="1" height="1" style="display:none"/>'
    msg.attach(MIMEText(corpo, "plain", "utf-8"))
    msg.attach(MIMEText(f"<html><body>{corpo_html}</body></html>", "html", "utf-8"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_PASS)
        server.sendmail(GMAIL_USER, destinatario, msg.as_string())

# ──────────────────────────────────────────────
# FLASK APP
# ──────────────────────────────────────────────

app = Flask(__name__)
disparo_status = {"rodando": False, "cancelar": False, "log": [], "progresso": 0, "total": 0}

@app.route("/")
def index():
    return render_template("index.html", gmail_user=GMAIL_USER, gmail_ok=bool(GMAIL_USER and GMAIL_PASS))

@app.route("/track/<tracking_id>")
def track(tracking_id):
    aberturas = get_aberturas()
    aberturas[tracking_id] = {"abertura": datetime.now().strftime("%d/%m/%Y %H:%M")}
    salvar(F_ABERTURAS, aberturas)
    pixel = bytes.fromhex("47494638396101000100800000ffffff0000002cffffff0000000000000002024401003b00")[:43]
    return pixel, 200, {"Content-Type": "image/gif"}

# ── API: CONTATOS ────────────────────────────

@app.route("/api/contatos", methods=["GET"])
def api_contatos():
    return jsonify(get_contatos())

@app.route("/api/contatos", methods=["POST"])
def api_add_contato():
    contatos = get_contatos()
    d = request.json
    if not d.get("nome") or not d.get("email") or not d.get("empresa"):
        return jsonify({"erro": "Campos obrigatórios faltando"}), 400
    contatos.append({
        "id": str(uuid.uuid4())[:8],
        "nome": d["nome"], "email": d["email"], "empresa": d["empresa"],
        "tipo": d.get("tipo", "clinica"), "whatsapp": d.get("whatsapp", "")
    })
    salvar(F_CONTATOS, contatos)
    return jsonify({"ok": True, "contatos": contatos})

@app.route("/api/contatos/<cid>", methods=["DELETE"])
def api_del_contato(cid):
    contatos = [c for c in get_contatos() if c.get("id") != cid]
    salvar(F_CONTATOS, contatos)
    return jsonify({"ok": True, "contatos": contatos})

@app.route("/api/contatos/limpar", methods=["POST"])
def api_limpar_contatos():
    salvar(F_CONTATOS, [])
    return jsonify({"ok": True})

@app.route("/api/contatos/importar", methods=["POST"])
def api_importar_contatos():
    file = request.files.get("arquivo")
    if not file:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    contatos = get_contatos()
    add = 0
    filename = file.filename.lower()

    if filename.endswith(".csv"):
        stream = io.StringIO(file.stream.read().decode("utf-8-sig"))
        reader = csv.DictReader(stream)
        for row in reader:
            nome = (row.get("nome") or row.get("Nome") or "").strip()
            email = (row.get("email") or row.get("Email") or row.get("E-mail") or "").strip()
            empresa = (row.get("empresa") or row.get("Empresa") or "").strip()
            tipo = (row.get("tipo") or row.get("Tipo") or "clinica").strip().lower()
            wpp = (row.get("whatsapp") or row.get("WhatsApp") or "").strip()
            if nome and email:
                contatos.append({"id": str(uuid.uuid4())[:8], "nome": nome, "email": email,
                                  "empresa": empresa, "tipo": tipo, "whatsapp": wpp})
                add += 1
    elif filename.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.stream.read()))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        aliases = {"nome": ["nome", "name"], "email": ["email", "e-mail", "mail"],
                   "empresa": ["empresa", "company"], "tipo": ["tipo", "type"],
                   "whatsapp": ["whatsapp", "telefone", "fone", "celular"]}
        def col(row, k):
            for a in aliases.get(k, [k]):
                if a in headers:
                    return str(row[headers.index(a)].value or "").strip()
            return ""
        for row in ws.iter_rows(min_row=2):
            nome, email = col(row, "nome"), col(row, "email")
            empresa = col(row, "empresa")
            tipo = col(row, "tipo").lower() or "clinica"
            wpp = col(row, "whatsapp")
            if nome and email:
                contatos.append({"id": str(uuid.uuid4())[:8], "nome": nome, "email": email,
                                  "empresa": empresa, "tipo": tipo, "whatsapp": wpp})
                add += 1

    salvar(F_CONTATOS, contatos)
    return jsonify({"ok": True, "adicionados": add, "contatos": contatos})

@app.route("/api/contatos/modelo")
def api_modelo_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["nome", "email", "empresa", "tipo", "whatsapp"])
    writer.writerow(["João Silva", "joao@clinica.com", "Clínica Bella", "clinica", "11999990001"])
    writer.writerow(["Maria Souza", "maria@contabil.com", "Contábil MP", "escritorio", "11999990002"])
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name="modelo_contatos.csv")

# ── API: TEMPLATES ───────────────────────────

@app.route("/api/templates", methods=["GET"])
def api_templates():
    return jsonify(get_templates())

@app.route("/api/templates", methods=["POST"])
def api_add_template():
    templates = get_templates()
    d = request.json
    novo = {"id": proximo_id(templates), "nome": d.get("nome", "Novo template"),
            "tipo": d.get("tipo", "clinica"), "assunto": d.get("assunto", ""),
            "corpo": d.get("corpo", "")}
    templates.append(novo)
    salvar(F_TEMPLATES, templates)
    return jsonify({"ok": True, "templates": templates, "novo_id": novo["id"]})

@app.route("/api/templates/<int:tid>", methods=["PUT"])
def api_edit_template(tid):
    templates = get_templates()
    d = request.json
    for t in templates:
        if t["id"] == tid:
            t.update({"nome": d.get("nome", t["nome"]), "tipo": d.get("tipo", t["tipo"]),
                       "assunto": d.get("assunto", t["assunto"]), "corpo": d.get("corpo", t["corpo"])})
    salvar(F_TEMPLATES, templates)
    return jsonify({"ok": True, "templates": templates})

@app.route("/api/templates/<int:tid>", methods=["DELETE"])
def api_del_template(tid):
    templates = [t for t in get_templates() if t["id"] != tid]
    salvar(F_TEMPLATES, templates)
    return jsonify({"ok": True, "templates": templates})

@app.route("/api/templates/<int:tid>/duplicar", methods=["POST"])
def api_dup_template(tid):
    templates = get_templates()
    orig = next((t for t in templates if t["id"] == tid), None)
    if not orig:
        return jsonify({"erro": "Não encontrado"}), 404
    copia = dict(orig)
    copia["id"] = proximo_id(templates)
    copia["nome"] = f"{orig['nome']} (cópia)"
    templates.append(copia)
    salvar(F_TEMPLATES, templates)
    return jsonify({"ok": True, "templates": templates})

# ── API: BLACKLIST ───────────────────────────

@app.route("/api/blacklist", methods=["GET"])
def api_blacklist():
    return jsonify(get_blacklist())

@app.route("/api/blacklist", methods=["POST"])
def api_add_blacklist():
    bl = get_blacklist()
    email = request.json.get("email", "").strip()
    if email and email not in bl:
        bl.append(email)
        salvar(F_BLACKLIST, bl)
    return jsonify({"ok": True, "blacklist": bl})

@app.route("/api/blacklist/<path:email>", methods=["DELETE"])
def api_del_blacklist(email):
    bl = [b for b in get_blacklist() if b != email]
    salvar(F_BLACKLIST, bl)
    return jsonify({"ok": True, "blacklist": bl})

# ── API: DISPARO ──────────────────────────────

@app.route("/api/disparar", methods=["POST"])
def api_disparar():
    global disparo_status
    d = request.json
    template_id = d.get("template_id")
    contato_ids = d.get("contato_ids", [])
    nome_campanha = d.get("campanha", f"Campanha {datetime.now().strftime('%d/%m/%Y')}")
    delay_min = int(d.get("delay_min", 5))
    delay_max = int(d.get("delay_max", 10))

    if not GMAIL_USER or not GMAIL_PASS:
        return jsonify({"erro": "Configure o .env"}), 400

    templates = get_templates()
    tmpl = next((t for t in templates if t["id"] == template_id), None)
    if not tmpl:
        return jsonify({"erro": "Template não encontrado"}), 400

    contatos = get_contatos()
    selecionados = [c for c in contatos if c["id"] in contato_ids]
    if not selecionados:
        return jsonify({"erro": "Nenhum contato selecionado"}), 400

    if disparo_status["rodando"]:
        return jsonify({"erro": "Já existe um disparo em andamento"}), 400

    disparo_status = {"rodando": True, "cancelar": False, "log": [], "progresso": 0, "total": len(selecionados)}

    def worker():
        global disparo_status
        blacklist = [b.lower() for b in get_blacklist()]
        historico = get_historico()
        followups = get_followup()
        campanhas = get_campanhas()
        ok = erro = bloqueado = 0
        camp_id = str(uuid.uuid4())[:8]
        registros = []

        for i, c in enumerate(selecionados):
            if disparo_status["cancelar"]:
                disparo_status["log"].append({"tipo": "cancelado", "msg": "Disparo cancelado pelo usuário"})
                break

            if c["email"].lower() in blacklist:
                disparo_status["log"].append({"tipo": "blacklist", "msg": f"Bloqueado: {c['nome']} ({c['email']})"})
                registros.append({"data": datetime.now().strftime("%d/%m/%Y %H:%M"), "nome": c["nome"],
                    "email": c["email"], "empresa": c["empresa"], "tipo": c["tipo"], "template": tmpl["nome"],
                    "campanha": nome_campanha, "status": "Blacklist", "erro": "", "tracking_id": ""})
                bloqueado += 1
                disparo_status["progresso"] = i + 1
                continue

            tid = str(uuid.uuid4())
            assunto = personalizar(tmpl["assunto"], c["nome"], c["empresa"])
            corpo = personalizar(tmpl["corpo"], c["nome"], c["empresa"])

            try:
                enviar_email(c["email"], assunto, corpo, tracking_id=tid)
                disparo_status["log"].append({"tipo": "sucesso", "msg": f"Enviado: {c['nome']} ({c['email']})"})
                registros.append({"data": datetime.now().strftime("%d/%m/%Y %H:%M"), "nome": c["nome"],
                    "email": c["email"], "empresa": c["empresa"], "tipo": c["tipo"], "template": tmpl["nome"],
                    "campanha": nome_campanha, "status": "Enviado", "erro": "", "tracking_id": tid})
                followups.append({"email": c["email"], "nome": c["nome"], "empresa": c["empresa"],
                    "tipo": c["tipo"], "template": tmpl["nome"], "data_envio": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "followup_enviado": False})
                ok += 1
            except Exception as ex:
                disparo_status["log"].append({"tipo": "erro", "msg": f"Erro: {c['email']} — {ex}"})
                registros.append({"data": datetime.now().strftime("%d/%m/%Y %H:%M"), "nome": c["nome"],
                    "email": c["email"], "empresa": c["empresa"], "tipo": c["tipo"], "template": tmpl["nome"],
                    "campanha": nome_campanha, "status": "Erro", "erro": str(ex), "tracking_id": ""})
                erro += 1

            disparo_status["progresso"] = i + 1
            if i < len(selecionados) - 1 and not disparo_status["cancelar"]:
                time.sleep(random.randint(delay_min, delay_max))

        salvar(F_FOLLOWUP, followups)
        campanhas.append({"id": camp_id, "nome": nome_campanha, "template": tmpl["nome"],
            "data": datetime.now().strftime("%d/%m/%Y %H:%M"), "enviados": ok, "erros": erro,
            "bloqueados": bloqueado, "total": len(selecionados)})
        salvar(F_CAMPANHAS, campanhas)
        historico.extend(registros)
        salvar(F_HISTORICO, historico)
        disparo_status["rodando"] = False
        disparo_status["log"].append({"tipo": "fim", "msg": f"Concluído: {ok} enviados, {bloqueado} bloqueados, {erro} erros"})

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/disparo/status")
def api_disparo_status():
    return jsonify(disparo_status)

@app.route("/api/disparo/cancelar", methods=["POST"])
def api_disparo_cancelar():
    disparo_status["cancelar"] = True
    return jsonify({"ok": True})

# ── API: CAMPANHAS / HISTÓRICO ───────────────

@app.route("/api/campanhas")
def api_campanhas():
    return jsonify(list(reversed(get_campanhas())))

@app.route("/api/historico")
def api_historico():
    hist = get_historico()
    aberturas = get_aberturas()
    for r in hist:
        tid = r.get("tracking_id", "")
        r["abertura"] = aberturas.get(tid, {}).get("abertura", "") if tid else ""
    return jsonify(list(reversed(hist)))

@app.route("/api/historico/limpar", methods=["POST"])
def api_limpar_historico():
    salvar(F_HISTORICO, [])
    return jsonify({"ok": True})

@app.route("/api/historico/exportar")
def api_exportar_historico():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    hist = get_historico()
    aberturas = get_aberturas()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"
    headers = ["Data", "Nome", "E-mail", "Empresa", "Tipo", "Campanha", "Template", "Status", "Abertura", "Erro"]
    hf = PatternFill("solid", fgColor="2D6CDF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center")
    for ri, r in enumerate(hist, 2):
        tid = r.get("tracking_id", "")
        ab = aberturas.get(tid, {}).get("abertura", "—") if tid else "—"
        vals = [r.get("data",""), r.get("nome",""), r.get("email",""), r.get("empresa",""),
                r.get("tipo",""), r.get("campanha",""), r.get("template",""), r.get("status",""), ab, r.get("erro","")]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=v)
            if r.get("status") == "Enviado": cell.fill = PatternFill("solid", fgColor="E1F5EE")
            elif r.get("status") == "Erro": cell.fill = PatternFill("solid", fgColor="FCEBEB")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 50)

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      as_attachment=True, download_name="relatorio_apolux.xlsx")

# ── API: FOLLOW-UP ───────────────────────────

@app.route("/api/followup")
def api_followup():
    return jsonify(get_followup())

@app.route("/api/followup/disparar", methods=["POST"])
def api_disparar_followup():
    d = request.json
    assunto = d.get("assunto", "")
    corpo = d.get("corpo", "")
    dias = int(d.get("dias", 3))

    if not GMAIL_USER or not GMAIL_PASS:
        return jsonify({"erro": "Configure o .env"}), 400

    followups = get_followup()
    pendentes = [f for f in followups if not f.get("followup_enviado")]

    def worker():
        ok = erro = 0
        for f in pendentes:
            try:
                dt = datetime.strptime(f["data_envio"], "%d/%m/%Y %H:%M")
                if (datetime.now() - dt).days < dias:
                    continue
            except: pass
            try:
                a = personalizar(assunto, f["nome"], f["empresa"])
                c = personalizar(corpo, f["nome"], f["empresa"])
                enviar_email(f["email"], a, c)
                f["followup_enviado"] = True
                ok += 1
            except: erro += 1
        salvar(F_FOLLOWUP, followups)

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"ok": True, "pendentes": len(pendentes)})

# ── API: AGENDAMENTO ─────────────────────────

@app.route("/api/agendados", methods=["GET"])
def api_agendados():
    return jsonify(get_agendados())

@app.route("/api/agendados", methods=["POST"])
def api_add_agendado():
    ags = get_agendados()
    d = request.json
    try:
        dt = datetime.strptime(f"{d['data']} {d['hora']}", "%d/%m/%Y %H:%M")
    except:
        return jsonify({"erro": "Data/hora inválida"}), 400
    if dt <= datetime.now():
        return jsonify({"erro": "Data deve ser no futuro"}), 400

    templates = get_templates()
    tmpl = next((t for t in templates if t["id"] == d.get("template_id")), None)
    if not tmpl:
        return jsonify({"erro": "Template não encontrado"}), 400

    ag = {"id": str(uuid.uuid4())[:8], "nome": d.get("nome", ""), "template_id": tmpl["id"],
          "template_nome": tmpl["nome"], "data_hora": dt.strftime("%d/%m/%Y %H:%M"),
          "destinatarios": d.get("destinatarios", "Todos"), "status": "Aguardando"}
    ags.append(ag)
    salvar(F_AGENDADOS, ags)
    return jsonify({"ok": True, "agendados": ags})

@app.route("/api/agendados/<aid>", methods=["DELETE"])
def api_del_agendado(aid):
    ags = [a for a in get_agendados() if a["id"] != aid]
    salvar(F_AGENDADOS, ags)
    return jsonify({"ok": True, "agendados": ags})

def verificar_agendados_loop():
    while True:
        try:
            agora = datetime.now()
            ags = get_agendados()
            alterou = False
            for ag in ags:
                if ag["status"] != "Aguardando":
                    continue
                try:
                    dt = datetime.strptime(ag["data_hora"], "%d/%m/%Y %H:%M")
                except:
                    continue
                if agora >= dt:
                    executar_agendado(ag)
                    ag["status"] = "Executado"
                    alterou = True
            if alterou:
                salvar(F_AGENDADOS, ags)
        except: pass
        time.sleep(30)

def executar_agendado(ag):
    templates = get_templates()
    tmpl = next((t for t in templates if t["id"] == ag["template_id"]), None)
    if not tmpl: return
    contatos = get_contatos()
    dest = ag.get("destinatarios", "Todos")
    if dest == "Só clínicas": cs = [c for c in contatos if c["tipo"] == "clinica"]
    elif dest == "Só escritórios": cs = [c for c in contatos if c["tipo"] == "escritorio"]
    else: cs = contatos
    blacklist = [b.lower() for b in get_blacklist()]
    historico = get_historico()
    for c in cs:
        if c["email"].lower() in blacklist: continue
        try:
            tid = str(uuid.uuid4())
            a = personalizar(tmpl["assunto"], c["nome"], c["empresa"])
            corp = personalizar(tmpl["corpo"], c["nome"], c["empresa"])
            enviar_email(c["email"], a, corp, tracking_id=tid)
            historico.append({"data": datetime.now().strftime("%d/%m/%Y %H:%M"), "nome": c["nome"],
                "email": c["email"], "empresa": c["empresa"], "tipo": c["tipo"], "template": tmpl["nome"],
                "campanha": f"[Agendado] {ag['nome']}", "status": "Enviado", "erro": "", "tracking_id": tid})
        except Exception as ex:
            historico.append({"data": datetime.now().strftime("%d/%m/%Y %H:%M"), "nome": c["nome"],
                "email": c["email"], "empresa": c["empresa"], "tipo": c["tipo"], "template": tmpl["nome"],
                "campanha": f"[Agendado] {ag['nome']}", "status": "Erro", "erro": str(ex), "tracking_id": ""})
        time.sleep(random.randint(5, 10))
    salvar(F_HISTORICO, historico)

# ── API: WHATSAPP ─────────────────────────────

@app.route("/api/whatsapp/link", methods=["POST"])
def api_whatsapp_link():
    import urllib.parse
    d = request.json
    tel = "".join(filter(str.isdigit, d.get("telefone", "")))
    if not tel.startswith("55"):
        tel = "55" + tel
    msg = personalizar(d.get("mensagem", ""), d.get("nome", ""), d.get("empresa", ""))
    url = f"https://wa.me/{tel}?text={urllib.parse.quote(msg)}"

    wpp = get_wpp()
    wpp.append({"data": datetime.now().strftime("%d/%m/%Y %H:%M"), "nome": d.get("nome",""),
                "telefone": d.get("telefone",""), "empresa": d.get("empresa",""), "status": "Aberto"})
    salvar(F_WPP, wpp)
    return jsonify({"ok": True, "url": url})

@app.route("/api/whatsapp/historico")
def api_whatsapp_historico():
    return jsonify(list(reversed(get_wpp())))

# ── API: DASHBOARD STATS ─────────────────────

@app.route("/api/stats")
def api_stats():
    hist = get_historico()
    campanhas = get_campanhas()
    contatos = get_contatos()
    aberturas = get_aberturas()
    blacklist = get_blacklist()

    enviados = sum(1 for r in hist if r.get("status") == "Enviado")
    taxa_abertura = round(len(aberturas) / enviados * 100) if enviados > 0 else 0

    return jsonify({
        "contatos": len(contatos),
        "enviados": enviados,
        "aberturas": len(aberturas),
        "taxa_abertura": taxa_abertura,
        "campanhas": len(campanhas),
        "blacklist": len(blacklist),
        "gmail_conectado": bool(GMAIL_USER and GMAIL_PASS),
        "gmail_user": GMAIL_USER
    })


if __name__ == "__main__":
    threading.Thread(target=verificar_agendados_loop, daemon=True).start()
    app.run(host="127.0.0.1", port=5050, debug=False)
